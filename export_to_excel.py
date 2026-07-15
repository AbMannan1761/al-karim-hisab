import os
import sys
import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Paths
WORKSPACE_DIR = r"e:\user\OneDrive - Bangladesh Telecommunication Regulatory Commission\ABM\Sunnah\AL karim hisab"
DATA_DIR = os.path.join(WORKSPACE_DIR, "data")
JSON_DIR = os.path.join(DATA_DIR, "json")
OUTPUT_EXCEL = os.path.join(WORKSPACE_DIR, "hisab_alkarim_digitized.xlsx")

def clean_num(val):
    if val is None:
        return 0.0
    val_str = str(val).strip().replace(",", "")
    if not val_str or val_str == "〃" or val_str == '"':
        return 0.0
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def load_data():
    # Load index data
    index_file = os.path.join(JSON_DIR, "index_data.json")
    index_map = {}
    index_list = []
    if os.path.exists(index_file):
        with open(index_file, "r", encoding="utf-8") as f:
            index_list = json.load(f)
            for entry in index_list:
                try:
                    p_num = int(entry.get("page", 0))
                    if p_num:
                        if p_num not in index_map:
                            index_map[p_num] = []
                        index_map[p_num].append(entry)
                except Exception:
                    pass

    debit_rows = []
    credit_rows = []
    
    # We will also keep track of debit/credit sums per ledger page
    page_sums = {}
    page_phones = {}

    print("Compiling JSON files...")
    for page_num in range(3, 111):
        left_ledger = 2 * page_num
        right_ledger = 2 * page_num + 1
        
        # 1. Left Page (Debit)
        left_file = os.path.join(JSON_DIR, f"page_{page_num}_left.json")
        if os.path.exists(left_file):
            with open(left_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            transcribed_name = str(data.get("party_name") or "").strip()
            phone = str(data.get("phone") or "").strip()
            ledger_page = str(data.get("ledger_page_number") or left_ledger).strip()
            if phone:
                page_phones[left_ledger] = phone
            
            # Find expected index info
            idx_infos = index_map.get(left_ledger, [])
            idx_names = ", ".join([x.get("party_name", "") for x in idx_infos])
            idx_address = ", ".join([x.get("address", "") for x in idx_infos])
            idx_notes = "; ".join([x.get("notes", "") for x in idx_infos if x.get("notes", "")])
            
            table = data.get("debit_table", [])
            for row in table:
                qty = clean_num(row.get("qty"))
                taka = clean_num(row.get("taka"))
                total = clean_num(row.get("total"))
                
                # If total is 0 but qty and rate exist, compute it
                if total == 0.0 and qty > 0 and taka > 0:
                    total = qty * taka
                    
                row_data = {
                    "PDF Page": page_num,
                    "Ledger Page": ledger_page,
                    "Transcribed Name": transcribed_name,
                    "Expected Client (Index)": idx_names if idx_names else None,
                    "Address (Index)": idx_address if idx_address else None,
                    "Notes (Index)": idx_notes if idx_notes else None,
                    "No": row.get("no", ""),
                    "Date": row.get("date", ""),
                    "Details (বিঃ কাঃ)": row.get("bi_ka", ""),
                    "Description": row.get("description", ""),
                    "Size": row.get("size", ""),
                    "Model": row.get("model", ""),
                    "PD": row.get("pd", ""),
                    "Bill No": row.get("bill", ""),
                    "Qty": qty if qty > 0 else row.get("qty", ""),
                    "Rate (Taka)": taka if taka > 0 else row.get("taka", ""),
                    "Total": total if total > 0 else row.get("total", ""),
                    "Remarks": row.get("remarks", "")
                }
                debit_rows.append(row_data)
                
                # Add to ledger page sum
                if left_ledger not in page_sums:
                    page_sums[left_ledger] = {"debit": 0.0, "credit": 0.0}
                page_sums[left_ledger]["debit"] += total

        # 2. Right Page (Credit)
        right_file = os.path.join(JSON_DIR, f"page_{page_num}_right.json")
        if os.path.exists(right_file):
            with open(right_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                
            transcribed_name = str(data.get("party_name") or "").strip()
            phone = str(data.get("phone") or "").strip()
            ledger_page = str(data.get("ledger_page_number") or right_ledger).strip()
            
            # Find expected index info
            idx_infos = index_map.get(left_ledger, []) # Often credit page is indexed by the main left page
            if not idx_infos:
                idx_infos = index_map.get(right_ledger, [])
                
            idx_names = ", ".join([x.get("party_name", "") for x in idx_infos])
            idx_address = ", ".join([x.get("address", "") for x in idx_infos])
            idx_notes = "; ".join([x.get("notes", "") for x in idx_infos if x.get("notes", "")])
            
            table = data.get("credit_table", [])
            for row in table:
                amount = clean_num(row.get("amount"))
                row_data = {
                    "PDF Page": page_num,
                    "Ledger Page": ledger_page,
                    "Transcribed Name": transcribed_name,
                    "Expected Client (Index)": idx_names if idx_names else None,
                    "Address (Index)": idx_address if idx_address else None,
                    "Notes (Index)": idx_notes if idx_notes else None,
                    "No": row.get("no", ""),
                    "Date": row.get("date", ""),
                    "Amount": amount if amount > 0 else row.get("amount", ""),
                    "Remarks": row.get("remarks", "")
                }
                credit_rows.append(row_data)
                
                # Add to ledger page sum (attribute to left ledger page since they are the same client)
                client_page = left_ledger
                if client_page not in page_sums:
                    page_sums[client_page] = {"debit": 0.0, "credit": 0.0}
                page_sums[client_page]["credit"] += amount

    # 3. Create Summary list based on Index entries
    summary_rows = []
    for idx_entry in index_list:
        p_num = int(idx_entry.get("page", 0))
        debit_sum = 0.0
        credit_sum = 0.0
        if p_num in page_sums:
            debit_sum = page_sums[p_num]["debit"]
            credit_sum = page_sums[p_num]["credit"]
            
        summary_rows.append({
            "Client No": idx_entry.get("no", ""),
            "Party Name": idx_entry.get("party_name", ""),
            "Address": idx_entry.get("address", ""),
            "Ledger Page": p_num,
            "PDF Page": p_num // 2 if p_num > 0 else "",
            "Total Debit (Sales)": debit_sum,
            "Total Credit (Payments)": credit_sum,
            "Outstanding Balance": debit_sum - credit_sum,
            "Notes (Index)": idx_entry.get("notes", ""),
            "Phone": page_phones.get(p_num, "")
        })

    return summary_rows, debit_rows, credit_rows

def export():
    summary_rows, debit_rows, credit_rows = load_data()
    
    # Exclude Phone from Client Summary sheet
    df_summary_excel = pd.DataFrame([{k: v for k, v in row.items() if k != "Phone"} for row in summary_rows])
    df_debit = pd.DataFrame(debit_rows)
    df_credit = pd.DataFrame(credit_rows)
    
    print(f"Exporting to {OUTPUT_EXCEL}...")
    
    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        df_summary_excel.to_excel(writer, sheet_name="Client Summary", index=False)
        df_debit.to_excel(writer, sheet_name="Debit Entries (Sales)", index=False)
        df_credit.to_excel(writer, sheet_name="Credit Entries (Payments)", index=False)
        
        # Access sheets for styling
        workbook = writer.book
        
        # Styles
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border_side = Side(border_style="thin", color="D3D3D3")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Create individual client sheets
        for client in summary_rows:
            client_no = str(client.get("Client No", "")).strip()
            client_name = client.get("Party Name", "")
            client_page = client.get("Ledger Page", 0)
            if not client_page:
                continue
                
            client_pdf_page = client_page // 2
            
            # Format sheet name: {client_no}. {sanitized_name} (max 31 chars)
            clean_name = "".join(c for c in client_name if c not in r"\/?:*[]")
            sheet_name = f"{client_no}. {clean_name}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
                
            # Filter debits and credits for this client's page
            client_debits = [r for r in debit_rows if r.get("PDF Page") == client_pdf_page]
            client_credits = [r for r in credit_rows if r.get("PDF Page") == client_pdf_page]
            
            sheet = workbook.create_sheet(title=sheet_name)
            
            # 1. Header Banner
            sheet.merge_cells("A1:Q1")
            banner_cell = sheet["A1"]
            banner_cell.value = "আল করিম হিসাব - গ্রাহক খতিয়ান (Al Karim Hisab - Client Ledger)"
            banner_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
            banner_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            banner_cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 40
            
            # 2. Metadata Block (Row 3-5) with merged cells to prevent column width issues
            sheet.merge_cells("A3:C3")
            sheet["A3"] = "গ্রাহকের নাম (Name):"
            sheet["A3"].font = Font(name=font_family, size=10, bold=True)
            sheet["A3"].alignment = Alignment(horizontal="left", vertical="center")
            
            sheet.merge_cells("D3:K3")
            sheet["D3"] = client_name
            sheet["D3"].font = Font(name=font_family, size=10)
            sheet["D3"].alignment = Alignment(horizontal="left", vertical="center")
            
            sheet.merge_cells("L3:N3")
            sheet["L3"] = "লেজার পৃষ্ঠা (Page):"
            sheet["L3"].font = Font(name=font_family, size=10, bold=True)
            sheet["L3"].alignment = Alignment(horizontal="right", vertical="center")
            
            sheet.merge_cells("O3:Q3")
            sheet["O3"] = client_page
            sheet["O3"].font = Font(name=font_family, size=10)
            sheet["O3"].alignment = Alignment(horizontal="left", vertical="center")
            
            sheet.merge_cells("A4:C4")
            sheet["A4"] = "ঠিকানা (Address):"
            sheet["A4"].font = Font(name=font_family, size=10, bold=True)
            sheet["A4"].alignment = Alignment(horizontal="left", vertical="center")
            
            sheet.merge_cells("D4:K4")
            sheet["D4"] = client.get("Address", "")
            sheet["D4"].font = Font(name=font_family, size=10)
            sheet["D4"].alignment = Alignment(horizontal="left", vertical="center")
            
            sheet.merge_cells("L4:N4")
            sheet["L4"] = "ফোন নম্বর (Phone):"
            sheet["L4"].font = Font(name=font_family, size=10, bold=True)
            sheet["L4"].alignment = Alignment(horizontal="right", vertical="center")
            
            sheet.merge_cells("O4:Q4")
            sheet["O4"] = client.get("Phone", "N/A") or "N/A"
            sheet["O4"].font = Font(name=font_family, size=10)
            sheet["O4"].alignment = Alignment(horizontal="left", vertical="center")
            
            sheet.merge_cells("A5:C5")
            sheet["A5"] = "মন্তব্য (Notes):"
            sheet["A5"].font = Font(name=font_family, size=10, bold=True)
            sheet["A5"].alignment = Alignment(horizontal="left", vertical="center")
            
            sheet.merge_cells("D5:K5")
            sheet["D5"] = client.get("Notes (Index)", "")
            sheet["D5"].font = Font(name=font_family, size=10)
            sheet["D5"].alignment = Alignment(horizontal="left", vertical="center")
            
            # 3. KPI Summary cards (Row 7-8)
            kpi_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
            kpi_border = Border(left=Side(style='thin', color='B0C4DE'),
                                right=Side(style='thin', color='B0C4DE'),
                                top=Side(style='thin', color='B0C4DE'),
                                bottom=Side(style='thin', color='B0C4DE'))
            
            # Card 1: Total Sales (A7:E8)
            sheet.merge_cells("A7:E7")
            sheet.merge_cells("A8:E8")
            sheet["A7"] = "সর্বমোট বিক্রয় (Total Purchases)"
            sheet["A7"].font = Font(name=font_family, size=9, color="555555")
            sheet["A7"].alignment = Alignment(horizontal="center")
            
            # Card 2: Total Payments (G7:K8)
            sheet.merge_cells("G7:K7")
            sheet.merge_cells("G8:K8")
            sheet["G7"] = "সর্বমোট আদায় (Total Payments)"
            sheet["G7"].font = Font(name=font_family, size=9, color="555555")
            sheet["G7"].alignment = Alignment(horizontal="center")
            
            # Card 3: Outstanding (M7:Q8)
            sheet.merge_cells("M7:Q7")
            sheet.merge_cells("M8:Q8")
            sheet["M7"] = "অবশিষ্ট বকেয়া (Outstanding)"
            sheet["M7"].font = Font(name=font_family, size=9, color="555555")
            sheet["M7"].alignment = Alignment(horizontal="center")
            
            for r in [7, 8]:
                for c in range(1, 6):
                    sheet.cell(row=r, column=c).fill = kpi_fill
                    sheet.cell(row=r, column=c).border = kpi_border
                for c in range(7, 12):
                    sheet.cell(row=r, column=c).fill = kpi_fill
                    sheet.cell(row=r, column=c).border = kpi_border
                for c in range(13, 18):
                    sheet.cell(row=r, column=c).fill = kpi_fill
                    sheet.cell(row=r, column=c).border = kpi_border
            
            # 4. Table Headers (Row 10)
            sheet.merge_cells("A10:L10")
            cell_debit_hdr = sheet["A10"]
            cell_debit_hdr.value = "ডেবিট এন্ট্রি সমূহ (Debit - Bills/Sales)"
            cell_debit_hdr.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            cell_debit_hdr.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell_debit_hdr.alignment = Alignment(horizontal="center", vertical="center")
            
            sheet.merge_cells("N10:Q10")
            cell_credit_hdr = sheet["N10"]
            cell_credit_hdr.value = "ক্রেডিট এন্ট্রি সমূহ (Credit - Payments)"
            cell_credit_hdr.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            cell_credit_hdr.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell_credit_hdr.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[10].height = 24
            
            # Column Headers (Row 11)
            debit_headers = ["No", "Date", "Details (বিঃ কাঃ)", "Description", "Size", "Model", "PD", "Bill No", "Qty", "Rate (Taka)", "Total", "Remarks"]
            credit_headers = ["No", "Date", "Amount", "Remarks"]
            
            for col_idx, text in enumerate(debit_headers, start=1):
                cell = sheet.cell(row=11, column=col_idx, value=text)
                cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="5C82AD", end_color="5C82AD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
            for col_idx, text in enumerate(credit_headers, start=14):
                cell = sheet.cell(row=11, column=col_idx, value=text)
                cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4E9F5D", end_color="4E9F5D", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.row_dimensions[11].height = 24
            
            # 5. Data Rows (Row 12+)
            max_rows = max(len(client_debits), len(client_credits))
            for i in range(max_rows):
                row_idx = 12 + i
                
                # Write Debit
                if i < len(client_debits):
                    row_data = client_debits[i]
                    debit_values = [
                        row_data.get("No", ""),
                        row_data.get("Date", ""),
                        row_data.get("Details (বিঃ কাঃ)", ""),
                        row_data.get("Description", ""),
                        row_data.get("Size", ""),
                        row_data.get("Model", ""),
                        row_data.get("PD", ""),
                        row_data.get("Bill No", ""),
                        row_data.get("Qty", ""),
                        row_data.get("Rate (Taka)", ""),
                        row_data.get("Total", ""),
                        row_data.get("Remarks", "")
                    ]
                    for col_idx, val in enumerate(debit_values, start=1):
                        cell = sheet.cell(row=row_idx, column=col_idx, value=val)
                        if col_idx in [9, 10, 11] and isinstance(val, (int, float)):
                            cell.number_format = "#,##0"
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            
                # Write Credit
                if i < len(client_credits):
                    row_data = client_credits[i]
                    credit_values = [
                        row_data.get("No", ""),
                        row_data.get("Date", ""),
                        row_data.get("Amount", ""),
                        row_data.get("Remarks", "")
                    ]
                    for col_idx, val in enumerate(credit_values, start=14):
                        cell = sheet.cell(row=row_idx, column=col_idx, value=val)
                        if col_idx == 16 and isinstance(val, (int, float)):
                            cell.number_format = "#,##0"
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            debit_end_row = 11 + len(client_debits)
            credit_end_row = 11 + len(client_credits)
            
            # Debit Total Row
            debit_total_row = debit_end_row + 1
            sheet.cell(row=debit_total_row, column=10, value="Total:").font = Font(name=font_family, size=10, bold=True)
            sheet.cell(row=debit_total_row, column=10).alignment = Alignment(horizontal="right")
            total_debit_cell = sheet.cell(row=debit_total_row, column=11, value=f"=SUM(K12:K{debit_end_row})")
            total_debit_cell.font = Font(name=font_family, size=10, bold=True)
            total_debit_cell.number_format = "#,##0"
            total_debit_cell.alignment = Alignment(horizontal="right")
            
            # Credit Total Row
            credit_total_row = credit_end_row + 1
            sheet.cell(row=credit_total_row, column=15, value="Total:").font = Font(name=font_family, size=10, bold=True)
            sheet.cell(row=credit_total_row, column=15).alignment = Alignment(horizontal="right")
            total_credit_cell = sheet.cell(row=credit_total_row, column=16, value=f"=SUM(P12:P{credit_end_row})")
            total_credit_cell.font = Font(name=font_family, size=10, bold=True)
            total_credit_cell.number_format = "#,##0"
            total_credit_cell.alignment = Alignment(horizontal="right")
            
            # Apply borders
            thin_border = Border(left=Side(style='thin', color='D3D3D3'),
                                 right=Side(style='thin', color='D3D3D3'),
                                 top=Side(style='thin', color='D3D3D3'),
                                 bottom=Side(style='thin', color='D3D3D3'))
            
            double_bottom_border = Border(left=Side(style='thin', color='D3D3D3'),
                                          right=Side(style='thin', color='D3D3D3'),
                                          top=Side(style='thin', color='D3D3D3'),
                                          bottom=Side(style='double', color='000000'))
                                          
            for r in range(11, debit_total_row + 1):
                for c in range(1, 13):
                    sheet.cell(row=r, column=c).border = thin_border
            sheet.cell(row=debit_total_row, column=11).border = double_bottom_border
            
            for r in range(11, credit_total_row + 1):
                for c in range(14, 18):
                    sheet.cell(row=r, column=c).border = thin_border
            sheet.cell(row=credit_total_row, column=16).border = double_bottom_border
            
            # 6. Set KPI Values referencing total cells
            sheet["A8"] = f"=K{debit_total_row}"
            sheet["A8"].font = Font(name=font_family, size=14, bold=True, color="366092")
            sheet["A8"].alignment = Alignment(horizontal="center")
            sheet["A8"].number_format = "#,##0"
            
            sheet["G8"] = f"=P{credit_total_row}"
            sheet["G8"].font = Font(name=font_family, size=14, bold=True, color="2E7D32")
            sheet["G8"].alignment = Alignment(horizontal="center")
            sheet["G8"].number_format = "#,##0"
            
            sheet["M8"] = "=A8-G8"
            sheet["M8"].font = Font(name=font_family, size=14, bold=True, color="C62828")
            sheet["M8"].alignment = Alignment(horizontal="center")
            sheet["M8"].number_format = "#,##0"
            
        # Format columns and layout for all sheets
        for name in workbook.sheetnames:
            sheet = workbook[name]
            
            if name in ["Client Summary", "Debit Entries (Sales)", "Credit Entries (Payments)"]:
                # Format header row
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Auto-fit column widths
                for col in sheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        cell.font = Font(name=font_family, size=10)
                        cell.border = cell_border
                        
                        # Formatting values
                        if cell.row > 1:
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal="right")
                                if name == "Client Summary" and cell.column in [6, 7, 8]:
                                    cell.number_format = "#,##0"
                                elif name == "Debit Entries (Sales)" and cell.column in [15, 16, 17]:
                                    cell.number_format = "#,##0"
                                elif name == "Credit Entries (Payments)" and cell.column in [9]:
                                    cell.number_format = "#,##0"
                            else:
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                                
                        val_str = str(cell.value or "")
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                    
                    sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
                sheet.row_dimensions[1].height = 28
            else:
                # Client-specific sheet auto-fit (row >= 10)
                for col in sheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.row >= 10:
                            val_str = str(cell.value or "")
                            if len(val_str) > max_len:
                                max_len = len(val_str)
                            if col_letter == "E":
                                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")
                    if col_letter == "E":
                        sheet.column_dimensions[col_letter].width = 6
                    else:
                        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    print("Excel file created successfully!")

if __name__ == "__main__":
    export()
